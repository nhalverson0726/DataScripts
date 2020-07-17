#! python3
#clLifetime.py copies the lifetime data of specified die name to
#Chlorine Lifetime.xlms

import openpyxl, os, xlwings
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series
)
from openpyxl.utils.cell import get_column_letter

def copyToSheet(die, test):
    
    dieSheets = []
    for sheet in range(len(clSheets)):
        if (
            test.lower() in clSheets[sheet].lower()
            and 'singlearray' not in clSheets[sheet].lower()
            and 'type' not in clSheets[sheet].lower()
            ):
            dieSheets.append(clSheets[sheet])
            
    if len(dieSheets) > 0:
        
        print('Copying ' + die + ' ' + test)
        ws = lifetime.create_sheet(title = die + ' ' +test)            
        pasteRow = 2
        for sheet in range(len(dieSheets)):
            active = clWings.sheets[dieSheets[sheet]]
            for i in range(1, 22):
                for j in range(1,3):
                    try:
                        ws.cell(row=i+pasteRow-1, column=j).value = (
                            active.range(get_column_letter(j)+str(i+32)).value
                            )
                    except:
                        print(get_column_letter(j) + i+pasteRow-1 + ' Not Filled')
            pasteRow = pasteRow + 21
            
        ws.cell(row = 1, column=1).value = '[Cl2] (ppm)'
        ws.cell(row = 1, column=2).value = 'Response (nA)'
        ws.cell(row = 1, column=3).value = 'Calibration (nA)'
        ws.cell(row = 1, column=5).value = 'Slope'
        ws.cell(row = 2, column=5).value = 'Intercept'
        ws.cell(row = 1, column=6).value = (
            clWings.sheets[dieSheets[0]].range('B12').value
            )
        ws.cell(row = 2, column=6).value = (
            clWings.sheets[dieSheets[0]].range('B13').value
            )

        for i in range(2, ws.max_row):
            try:
                ws.cell(row=i, column=3).value = (
                float(ws['F1'].value)
                * float(ws.cell(row=i, column=1).value)
                + float(ws['F2'].value)
                )
            except:
                print('Row ' + str(i) + ' Not Filled')
            
        chart = ScatterChart()
        chart.title = str(die + ' ' + test + ' Lifetime (' + str(len(dieSheets)) + 'Calibrations)')
        chart.x_axis.title = '[Cl2] (ppm)'
        chart.y_axis.title = 'Response (nA)'

        xvalues = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        roamValues = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        roamSeries = Series(roamValues, xvalues, title_from_data=True)
        calValues =  Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
        calSeries = Series(calValues, xvalues, title_from_data=True, )
        roamSeries.marker=openpyxl.chart.marker.Marker('x')
        roamSeries.graphicalProperties.line.noFill=True
        chart.series.append(roamSeries)
        chart.series.append(calSeries)

        cs.add_chart(chart, 'A' + str(len(lifetime.worksheets) * 15 - 29))

print('Opening Files...')
os.chdir(
    r'C:\Users\e-sens\Google Drive\Lab Data\Sensor Testing\Results\Sensor Functionality\Lifetime')
lifetime = openpyxl.Workbook()
cs = lifetime.active
cs.title = 'Chart Sheet'
os.chdir(r'C:\Users\e-sens\Documents')
dies = ['Chip_082019_1', 'Chip_082119_1', 'Chip_091119_1', 'Chip_091119_2',
       'Chip_100719_Gold', 'Chip_601_0002', 'Chip_602_0002', 'Chip_602_002'
       ]
test = ['fcl', 'tcl']
        
for i in range(len(dies)):
    cl = load_workbook(filename =dies[i] + '.xlsx')
    clWings = xlwings.Book(dies[i] + '.xlsx')
    clSheets = cl.sheetnames
    for j in range(len(test)):
        copyToSheet(dies[i], test[j])
    
    
lifetime.save('Chlorine Lifetime 102819 Jason.xlsx')    
print('Done.')
